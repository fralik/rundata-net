#!/usr/bin/env python3
"""Import image links from Excel into runes_imagelink.

Expected columns:
  col1: inscription ID
  col2: image URL
  col3: caption/info (shown under avatar)
"""

from __future__ import annotations

import argparse
import re
import sqlite3
from collections import Counter
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl import load_workbook


PAREN_RE = re.compile(r"\(([^()]*)\)")
NIER_RE = re.compile(r"^NI[æÆ]R\s+(\d+)\s*$")


def id_candidates(raw_id: str) -> List[str]:
    s = raw_id.strip()
    out: List[str] = []
    seen = set()

    def add(v: str) -> None:
        v = v.strip()
        if v and v not in seen:
            seen.add(v)
            out.append(v)

    add(s)

    m = re.search(r"\(([^()]*)\)\s*$", s)
    if m:
        add(m.group(1))

    for g in PAREN_RE.findall(s):
        add(g)

    # Match previous import behavior for NIæR aliases.
    for base in list(out):
        m2 = NIER_RE.match(base)
        if m2:
            add(f"N {m2.group(1)}")

    return out


def ensure_info_column(conn: sqlite3.Connection) -> None:
    cols = [r[1] for r in conn.execute("PRAGMA table_info('runes_imagelink')")]
    if "info" not in cols:
        conn.execute("ALTER TABLE runes_imagelink ADD COLUMN info varchar(1500) NOT NULL DEFAULT ''")
        conn.commit()


def load_signature_maps(conn: sqlite3.Connection) -> Tuple[Dict[str, Tuple[int, Optional[int]]], Dict[int, int]]:
    sig_map = {
        r["signature_text"]: (r["id"], r["parent_id"])
        for r in conn.execute("SELECT id, signature_text, parent_id FROM signatures")
    }
    meta_map = {r["signature_id"]: r["id"] for r in conn.execute("SELECT id, signature_id FROM meta_information")}
    return sig_map, meta_map


def resolve_meta_id(raw_id: str, sig_map: Dict[str, Tuple[int, Optional[int]]], meta_map: Dict[int, int]) -> Optional[int]:
    for cand in id_candidates(raw_id):
        hit = sig_map.get(cand)
        if hit is None:
            continue
        sig_id, parent_id = hit
        if sig_id in meta_map:
            return meta_map[sig_id]
        if parent_id is not None and parent_id in meta_map:
            return meta_map[parent_id]
    return None


def run(db_path: Path, xlsx_paths: List[Path]) -> None:
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row

    ensure_info_column(conn)
    sig_map, meta_map = load_signature_maps(conn)

    rows_total = 0
    matched_rows = 0
    inserted = 0
    updated = 0
    skipped_invalid = 0
    unmatched_counter: Counter[str] = Counter()

    for xlsx_path in xlsx_paths:
        wb = load_workbook(xlsx_path, read_only=True, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows(min_row=2, values_only=True):
                raw_id = "" if row[0] is None else str(row[0]).strip()
                link = "" if row[1] is None else str(row[1]).strip()
                info = "" if row[2] is None else str(row[2]).strip()

                if not raw_id:
                    continue
                rows_total += 1

                if not link or not (link.startswith("http://") or link.startswith("https://")):
                    skipped_invalid += 1
                    continue

                meta_id = resolve_meta_id(raw_id, sig_map, meta_map)
                if meta_id is None:
                    unmatched_counter[raw_id] += 1
                    continue

                matched_rows += 1

                # Prefer existing row for this exact image URL on this inscription.
                existing = conn.execute(
                    """
                    SELECT id FROM runes_imagelink
                    WHERE meta_id = ? AND (direct_url = ? OR link_url = ?)
                    ORDER BY id
                    LIMIT 1
                    """,
                    (meta_id, link, link),
                ).fetchone()

                info_value = info[:1500]
                if existing is None:
                    conn.execute(
                        "INSERT INTO runes_imagelink (link_url, direct_url, meta_id, info) VALUES (?, ?, ?, ?)",
                        (link, link, meta_id, info_value),
                    )
                    inserted += 1
                else:
                    conn.execute(
                        "UPDATE runes_imagelink SET link_url = ?, direct_url = ?, info = ? WHERE id = ?",
                        (link, link, info_value, existing["id"]),
                    )
                    updated += 1

        wb.close()

    conn.commit()

    print("IMPORT_IMAGES_REPORT")
    print(f"rows_total_with_id={rows_total}")
    print(f"matched_rows={matched_rows}")
    print(f"inserted_rows={inserted}")
    print(f"updated_rows={updated}")
    print(f"skipped_invalid_link_rows={skipped_invalid}")
    print(f"unmatched_rows={sum(unmatched_counter.values())}")
    print(f"unmatched_unique_ids={len(unmatched_counter)}")
    print("UNMATCHED_IDS_START")
    for k, v in unmatched_counter.most_common():
        print(f"{k}\t({v} rows)")
    print("UNMATCHED_IDS_END")

    conn.close()


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("--db", required=True)
    parser.add_argument("--xlsx", required=True, nargs="+")
    return parser.parse_args()


if __name__ == "__main__":
    args = parse_args()
    run(Path(args.db), [Path(p) for p in args.xlsx])
